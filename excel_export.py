"""Exportación a Excel con índice, hojas agrupadas y formato profesional.

Este módulo conserva el estilo del Excel final original: índice de navegación,
títulos fusionados, formatos de porcentaje, bordes, filtros y enlaces internos.
"""
import numpy as np
import pandas as pd
from config import REPORTS, CONVERSION_COLS
from eda import tabla_conversion


def limpiar_nombre_hoja(nombre):
    """
    Limpia nombres de hojas para que Excel no dé error.
    Excel permite máximo 31 caracteres y no permite ciertos símbolos.
    """
    caracteres_invalidos = ["\\", "/", "*", "?", ":", "[", "]"]
    for char in caracteres_invalidos:
        nombre = nombre.replace(char, "")
    return nombre[:31]

def preparar_tabla_excel(df):
    """
    Copia una tabla antes de escribirla en Excel.
    Evita modificar el DataFrame original.
    """
    if df is None:
        return pd.DataFrame()

    tabla = df.copy()

    # Si hay índices con información, los dejamos como columnas
    if tabla.index.name is not None:
        tabla = tabla.reset_index()

    return tabla

def traducir_columnas(df):
    """
    Traduce nombres de columnas frecuentes a español para que el Excel final
    sea más legible.
    """
    traducciones = {
        "modelo": "Modelo",
        "metrica": "Métrica",
        "valor": "Valor",
        "roc_auc": "Área bajo curva ROC",
        "average_precision": "Precisión media",
        "accuracy": "Exactitud",
        "precision_clase_1": "Precisión clase positiva",
        "recall_clase_1": "Sensibilidad clase positiva",
        "f1_clase_1": "F1 clase positiva",
        "threshold": "Umbral",
        "umbral": "Umbral",
        "precision": "Precisión",
        "recall": "Sensibilidad",
        "f1": "F1",
        "fp": "Falsos positivos",
        "fn": "Falsos negativos",
        "tp": "Verdaderos positivos",
        "tn": "Verdaderos negativos",
        "variable": "Variable",
        "Coef.": "Coeficiente",
        "Std.Err.": "Error estándar",
        "z": "Estadístico z",
        "P>|z|": "P-valor",
        "[0.025": "IC 2,5%",
        "0.975]": "IC 97,5%",
        "odds_ratio": "Odds ratio",
        "or_ci_2.5%": "OR IC 2,5%",
        "or_ci_97.5%": "OR IC 97,5%",
        "interpretacion": "Interpretación",
        "significatividad": "Significatividad",
        "clientes": "Clientes",
        "suscriptores": "Suscriptores",
        "tasa_conversion": "Tasa de conversión",
        "conversion_pct": "Tasa de conversión",
        "prob_media": "Probabilidad media",
        "prob_min": "Probabilidad mínima",
        "prob_max": "Probabilidad máxima",
        "pct_suscriptores": "% de suscriptores",
        "gain_acumulado": "Ganancia acumulada",
        "lift": "Lift",
        "lift_acumulado": "Lift acumulado",
        "decil": "Decil",
        "clientes_total": "Clientes totales",
        "clientes_primer_decil": "Clientes primer decil",
        "clientes_sin_primer_decil": "Clientes sin primer decil",
        "clientes_eliminados": "Clientes eliminados",
        "tasa_total": "Tasa total",
        "tasa_primer_decil": "Tasa primer decil",
        "tasa_sin_primer_decil": "Tasa sin primer decil",
        "probabilidad_media": "Probabilidad media",
        "pseudo_r2": "Pseudo R²",
        "log_likelihood": "Log-Likelihood",
        "aic": "AIC",
        "bic": "BIC",
        "n_observaciones": "Nº observaciones",
        "grados_libertad_modelo": "Grados libertad modelo",
        "convergencia": "Convergencia",
        "test": "Test",
        "p_value": "P-valor",
        "estadistico": "Estadístico",
        "asimetria": "Asimetría",
        "curtosis": "Curtosis",
        "resultado": "Resultado",
    }

    tabla = df.copy()
    tabla = tabla.rename(columns={col: traducciones.get(col, col) for col in tabla.columns})
    return tabla

def normalizar_porcentajes(df):
    """
    Convierte columnas que representan porcentajes a escala 0-1 para que Excel
    pueda mostrarlas con formato porcentual.

    Ejemplo:
    31.34 pasa a 0.3134 y Excel lo muestra como 31,34%.
    0.3134 se mantiene igual y Excel lo muestra como 31,34%.
    """
    tabla = df.copy()

    palabras_porcentaje = [
        "tasa",
        "porcentaje",
        "%",
        "precision",
        "precisión",
        "sensibilidad",
        "recall",
        "f1",
        "exactitud",
        "accuracy",
        "conversion",
        "conversión",
        "probabilidad",
        "ganancia acumulada",
    ]

    for col in tabla.columns:
        col_lower = str(col).lower()

        if any(palabra in col_lower for palabra in palabras_porcentaje):
            if pd.api.types.is_numeric_dtype(tabla[col]):
                max_val = tabla[col].max(skipna=True)

                # Si viene como 31.34, lo pasamos a 0.3134
                if pd.notna(max_val) and max_val > 1.5:
                    tabla[col] = tabla[col] / 100

    return tabla

def descripciones_hojas_excel():
    """
    Diccionario con la descripción de las hojas principales del Excel final.
    Si alguna hoja no aparece aquí, se le asignará una descripción genérica.
    """
    return {
        "01_resumen_modelos": "Resumen comparativo de los principales modelos: Logit, regresión logística y Random Forest.",
        "02_logit_final": "Detalle del modelo Logit seleccionado, incluyendo métricas del modelo, coeficientes, p-valores y odds ratios.",
        "03_odds_ratios": "Interpretación de los odds ratios del modelo seleccionado.",
        "04_multicolinealidad": "Diagnóstico de multicolinealidad mediante VIF para detectar variables redundantes.",
        "05_umbrales": "Comparación de distintos umbrales de decisión y su impacto en precisión, sensibilidad y F1.",
        "06_deciles": "Tablas de deciles y lift para priorizar clientes según probabilidad estimada.",
        "07_primer_decil": "Análisis específico del primer decil, es decir, el grupo con mayor probabilidad estimada.",
        "08_sin_primer_decil": "Análisis de la muestra restante tras excluir el primer decil.",
        "09_validacion_temporal": "Validación temporal del modelo usando datos antiguos para entrenar y datos recientes para evaluar.",
        "10_validacion_cruzada": "Resumen de la validación cruzada del modelo.",
        "11_normalidad": "Tests de normalidad aplicados a variables numéricas del modelo.",
        "01_descriptivos": "Resumen descriptivo general de las variables del dataset.",
        "conv_job": "Tasa de conversión por tipo de ocupación.",
        "conv_education": "Tasa de conversión por nivel educativo.",
        "conv_marital": "Tasa de conversión por estado civil.",
        "conv_contact": "Tasa de conversión por canal de contacto.",
        "conv_poutcome": "Tasa de conversión según resultado de campaña anterior.",
        "conv_income_group": "Tasa de conversión por grupo de ingresos.",
        "conv_age_group": "Tasa de conversión por grupo de edad.",
    }

def escribir_tabla_excel(writer, df, sheet_name, titulo=None, startrow=0, startcol=0):
    """
    Escribe una tabla en Excel usando openpyxl con formato mejorado.

    Mejora respecto a la versión anterior:
    - Fusiona el título a lo ancho de la tabla.
    - Ajusta columnas.
    - Formatea porcentajes.
    - Añade enlace de vuelta al índice.
    - Evita que el título quede cortado en la celda A1.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    sheet_name = limpiar_nombre_hoja(sheet_name)

    tabla = preparar_tabla_excel(df)
    tabla = traducir_columnas(tabla)
    tabla = normalizar_porcentajes(tabla)

    if titulo is None:
        titulo = sheet_name

    # Crear hoja si no existe
    if sheet_name not in writer.book.sheetnames:
        worksheet = writer.book.create_sheet(sheet_name)
    else:
        worksheet = writer.book[sheet_name]

    writer.sheets[sheet_name] = worksheet

    # Estilos generales
    color_azul = "1F4E78"
    color_azul_claro = "D9EAF7"
    color_gris = "F2F2F2"

    header_fill = PatternFill("solid", fgColor=color_azul_claro)
    title_fill = PatternFill("solid", fgColor=color_azul)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Determinamos ancho de la tabla
    n_cols = max(len(tabla.columns), 4)
    first_col = startcol + 1
    last_col = startcol + n_cols

    title_row = startrow + 1
    link_row = startrow + 2
    header_row = startrow + 4
    first_data_row = startrow + 5
    last_row = header_row + len(tabla)

    # Título fusionado
    worksheet.merge_cells(
        start_row=title_row,
        start_column=first_col,
        end_row=title_row,
        end_column=last_col
    )

    celda_titulo = worksheet.cell(row=title_row, column=first_col)
    celda_titulo.value = titulo
    celda_titulo.font = Font(bold=True, size=14, color="FFFFFF")
    celda_titulo.fill = title_fill
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    worksheet.row_dimensions[title_row].height = 24

    # Enlace de vuelta al índice
    if sheet_name != "00_indice":
        celda_link = worksheet.cell(row=link_row, column=first_col)
        celda_link.value = "Volver al índice"
        celda_link.hyperlink = "#'00_indice'!A1"
        celda_link.style = "Hyperlink"
        celda_link.font = Font(color="0563C1", underline="single")
        worksheet.row_dimensions[link_row].height = 18

    # Si no hay datos
    if tabla.empty:
        celda = worksheet.cell(row=header_row, column=first_col)
        celda.value = "No hay datos disponibles."
        celda.fill = PatternFill("solid", fgColor=color_gris)
        celda.border = border
        return startrow + 7

    # Escribimos la tabla con pandas
    tabla.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=header_row - 1,
        startcol=startcol,
        index=False
    )

    # Reasignamos por seguridad
    worksheet = writer.book[sheet_name]
    writer.sheets[sheet_name] = worksheet

    last_col = startcol + len(tabla.columns)
    last_row = header_row + len(tabla)

    # Cabecera
    for col in range(first_col, last_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[header_row].height = 30

    # Datos
    for row in range(first_data_row, last_row + 1):
        for col in range(first_col, last_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            nombre_columna = str(worksheet.cell(row=header_row, column=col).value).lower()

            if isinstance(cell.value, (int, float)):
                if any(p in nombre_columna for p in [
                    "tasa",
                    "%",
                    "precisión",
                    "sensibilidad",
                    "f1",
                    "exactitud",
                    "probabilidad",
                    "ganancia"
                ]):
                    cell.number_format = "0.00%"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.0000"

    # Ajuste de columnas
    for col_idx in range(first_col, last_col + 1):
        col_letter = get_column_letter(col_idx)

        valores = [
            worksheet.cell(row=row, column=col_idx).value
            for row in range(title_row, last_row + 1)
        ]

        max_len = max(len(str(v)) if v is not None else 0 for v in valores)

        # Evitamos columnas excesivamente estrechas o excesivamente anchas
        ancho = min(max(max_len + 2, 12), 38)
        worksheet.column_dimensions[col_letter].width = ancho

    # Damos más anchura a la primera columna si contiene texto
    worksheet.column_dimensions[get_column_letter(first_col)].width = max(
        worksheet.column_dimensions[get_column_letter(first_col)].width,
        22
    )

    # Filtro
    worksheet.auto_filter.ref = (
        f"{get_column_letter(first_col)}{header_row}:"
        f"{get_column_letter(last_col)}{last_row}"
    )

    # Congelar paneles solo si la tabla empieza al principio de la hoja
    if startrow == 0:
        worksheet.freeze_panes = worksheet.cell(row=first_data_row, column=first_col)

    return startrow + len(tabla) + 7

def crear_hoja_indice(writer, nombre_archivo="Excel de resultados"):
    """
    Crea una hoja inicial con enlaces internos al resto de hojas.
    Permite navegar sin tener que pulsar pestañas manualmente.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = writer.book
    descripciones = descripciones_hojas_excel()

    # Si ya existe, la eliminamos para reconstruirla
    if "00_indice" in wb.sheetnames:
        hoja_antigua = wb["00_indice"]
        wb.remove(hoja_antigua)

    ws = wb.create_sheet("00_indice", 0)

    # Estilos
    azul = "1F4E78"
    azul_claro = "D9EAF7"
    gris = "F2F2F2"

    title_fill = PatternFill("solid", fgColor=azul)
    header_fill = PatternFill("solid", fgColor=azul_claro)
    grey_fill = PatternFill("solid", fgColor=gris)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título principal
    ws.merge_cells("A1:D1")
    ws["A1"] = nombre_archivo
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Explicación
    ws.merge_cells("A3:D3")
    ws["A3"] = "Esta hoja permite navegar por el archivo y resume el contenido de cada pestaña."
    ws["A3"].font = Font(italic=True, color="404040")
    ws["A3"].alignment = Alignment(wrap_text=True)

    # Cabeceras
    headers = ["Nº", "Hoja", "Contenido", "Enlace"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas del índice
    fila = 6
    contador = 1

    for sheet in wb.worksheets:
        if sheet.title == "00_indice":
            continue

        nombre_hoja = sheet.title
        descripcion = descripciones.get(
            nombre_hoja,
            "Hoja de resultados generada durante el análisis."
        )

        ws.cell(row=fila, column=1).value = contador
        ws.cell(row=fila, column=2).value = nombre_hoja
        ws.cell(row=fila, column=3).value = descripcion
        ws.cell(row=fila, column=4).value = "Ir a hoja"

        # Hipervínculo interno
        ws.cell(row=fila, column=4).hyperlink = f"#'{nombre_hoja}'!A1"
        ws.cell(row=fila, column=4).style = "Hyperlink"
        ws.cell(row=fila, column=4).font = Font(color="0563C1", underline="single")

        for col in range(1, 5):
            cell = ws.cell(row=fila, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if contador % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=fila, column=col).fill = grey_fill

        fila += 1
        contador += 1

    # Anchos
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 75
    ws.column_dimensions["D"].width = 18

    # Altura de filas
    for row in range(6, fila):
        ws.row_dimensions[row].height = 34

    # Filtro
    ws.auto_filter.ref = f"A5:D{fila - 1}"

    # Congelar cabecera
    ws.freeze_panes = "A6"

    # Quitamos la hoja vacía por defecto si existe
    if "Sheet" in wb.sheetnames and len(wb["Sheet"]["A"]) == 1 and wb["Sheet"]["A1"].value is None:
        wb.remove(wb["Sheet"])

def tabla_comparacion_logit_inferencia(
    resumen_statsmodels,
    resumen_statsmodels_reducido,
    resumen_logit_red_sin_cp,
    resumen_logit_red_sin_cp_robusto
):
    """
    Agrupa los principales modelos Logit en una sola tabla comparativa.
    Compatible con resúmenes en formato elemento/valor o metrica/valor.
    """
    modelos = [
        ("Logit completo", resumen_statsmodels),
        ("Logit reducido", resumen_statsmodels_reducido),
        ("Logit reducido sin contacto previo", resumen_logit_red_sin_cp),
        ("Logit reducido sin contacto previo robusto", resumen_logit_red_sin_cp_robusto),
    ]

    filas = []

    for nombre, resumen in modelos:
        if resumen is None or len(resumen) == 0:
            continue

        temp = resumen.copy()

        if "elemento" in temp.columns and "valor" in temp.columns:
            dic = dict(zip(temp["elemento"], temp["valor"]))
        elif "metrica" in temp.columns and "valor" in temp.columns:
            dic = dict(zip(temp["metrica"], temp["valor"]))
        elif "Métrica" in temp.columns and "Valor" in temp.columns:
            dic = dict(zip(temp["Métrica"], temp["Valor"]))
        elif "Elemento" in temp.columns and "Valor" in temp.columns:
            dic = dict(zip(temp["Elemento"], temp["Valor"]))
        else:
            continue

        filas.append({
            "Modelo": nombre,
            "Pseudo R²": dic.get("pseudo_r2", dic.get("Pseudo R²", np.nan)),
            "Log-Likelihood": dic.get("log_likelihood", dic.get("Log-Likelihood", np.nan)),
            "AIC": dic.get("aic", dic.get("AIC", np.nan)),
            "BIC": dic.get("bic", dic.get("BIC", np.nan)),
            "Nº observaciones": dic.get("observaciones", dic.get("n_observaciones", dic.get("Nº observaciones", np.nan))),
            "Convergencia": dic.get("convergencia", dic.get("Convergencia", np.nan)),
        })

    return pd.DataFrame(filas)

def crear_excel_final_limpio(
    merged: pd.DataFrame,
    resultados_predictivos: dict[str, object],
    logit_final: dict[str, object],
    logit_robusto: dict[str, object],
    correlacion: pd.DataFrame,
    vif: pd.DataFrame,
    odds: pd.DataFrame,
    validacion_temp: dict[str, pd.DataFrame] | None,
    normalidad: pd.DataFrame,
) -> None:
    """Crea el Excel final manteniendo índice, pestañas agrupadas y formatos.

    A diferencia de un simple ``to_excel``, esta función escribe varias tablas
    relacionadas dentro de cada hoja, aplica formato, crea enlaces internos y
    añade una hoja índice al principio del libro.
    """
    ruta = REPORTS / "resultados_modelos_finales.xlsx"

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        # 01. Resumen de modelos
        fila = 0
        fila = escribir_tabla_excel(
            writer,
            resultados_predictivos.get("comparacion", pd.DataFrame()),
            "01_resumen_modelos",
            "Comparación de modelos predictivos",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            resultados_predictivos.get("metricas_base", pd.DataFrame()),
            "01_resumen_modelos",
            "Métricas del modelo base",
            startrow=fila,
        )

        # 02. Modelo Logit final
        fila = 0
        fila = escribir_tabla_excel(
            writer,
            logit_final.get("resumen", pd.DataFrame()),
            "02_logit_final",
            "Resumen del modelo Logit reducido sin contacto_previo",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            logit_final.get("tabla", pd.DataFrame()),
            "02_logit_final",
            "Coeficientes y odds ratios del modelo Logit final",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            logit_robusto.get("resumen", pd.DataFrame()),
            "02_logit_final",
            "Resumen del modelo robusto HC3",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            logit_robusto.get("tabla", pd.DataFrame()),
            "02_logit_final",
            "Coeficientes y odds ratios del modelo robusto HC3",
            startrow=fila,
        )

        # 03. Odds ratios
        escribir_tabla_excel(
            writer,
            odds,
            "03_odds_ratios",
            "Interpretación de odds ratios del modelo seleccionado",
        )

        # 04. Multicolinealidad
        fila = 0
        fila = escribir_tabla_excel(
            writer,
            vif,
            "04_multicolinealidad",
            "Diagnóstico de multicolinealidad mediante VIF",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            correlacion.reset_index().rename(columns={"index": "variable"}),
            "04_multicolinealidad",
            "Matriz de correlación entre variables numéricas",
            startrow=fila,
        )

        # 05. Umbrales
        fila = 0
        fila = escribir_tabla_excel(
            writer,
            resultados_predictivos.get("umbrales_base", pd.DataFrame()),
            "05_umbrales",
            "Análisis de umbrales - modelo base",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            resultados_predictivos.get("umbrales_balanceado", pd.DataFrame()),
            "05_umbrales",
            "Análisis de umbrales - modelo balanceado",
            startrow=fila,
        )

        # 06. Deciles
        fila = 0
        fila = escribir_tabla_excel(
            writer,
            resultados_predictivos.get("deciles_base", pd.DataFrame()),
            "06_deciles",
            "Tabla de deciles - modelo base fuera de muestra",
            startrow=fila,
        )
        fila = escribir_tabla_excel(
            writer,
            resultados_predictivos.get("deciles_balanceado", pd.DataFrame()),
            "06_deciles",
            "Tabla de deciles - modelo balanceado fuera de muestra",
            startrow=fila,
        )



        # 07. Primer decil
        deciles_base = resultados_predictivos.get("deciles_base", pd.DataFrame())
        if not deciles_base.empty and "decil" in deciles_base.columns:
            primer_decil = deciles_base[deciles_base["decil"] == 1].copy()
        else:
            primer_decil = pd.DataFrame()
        escribir_tabla_excel(
            writer,
            primer_decil,
            "07_primer_decil",
            "Resumen del primer decil fuera de muestra",
        )

        # 08. Sin primer decil
        if not deciles_base.empty and "decil" in deciles_base.columns:
            sin_primer = deciles_base[deciles_base["decil"] != 1].copy()
            resumen_sin_primer = pd.DataFrame({
                "elemento": [
                    "clientes_sin_primer_decil",
                    "suscriptores_sin_primer_decil",
                    "tasa_conversion_sin_primer_decil",
                    "nota",
                ],
                "valor": [
                    sin_primer["clientes"].sum(),
                    sin_primer["suscriptores"].sum(),
                    sin_primer["suscriptores"].sum() / sin_primer["clientes"].sum(),
                    "Resumen calculado sobre los deciles fuera de muestra del modelo base.",
                ],
            })
        else:
            resumen_sin_primer = pd.DataFrame()
        escribir_tabla_excel(
            writer,
            resumen_sin_primer,
            "08_sin_primer_decil",
            "Resumen de la muestra sin primer decil",
        )

        # 09. Validación temporal
        fila = 0
        if validacion_temp is not None:
            fila = escribir_tabla_excel(
                writer,
                validacion_temp.get("info", pd.DataFrame()),
                "09_validacion_temporal",
                "Información de la validación temporal",
                startrow=fila,
            )
            fila = escribir_tabla_excel(
                writer,
                validacion_temp.get("metricas", pd.DataFrame()),
                "09_validacion_temporal",
                "Métricas de la validación temporal",
                startrow=fila,
            )
            fila = escribir_tabla_excel(
                writer,
                validacion_temp.get("matriz", pd.DataFrame()),
                "09_validacion_temporal",
                "Matriz de confusión de la validación temporal",
                startrow=fila,
            )

        # 10. Validación cruzada
        escribir_tabla_excel(
            writer,
            resultados_predictivos.get("cv_resumen", pd.DataFrame()),
            "10_validacion_cruzada",
            "Validación cruzada - resumen",
        )

        # 11. Normalidad
        escribir_tabla_excel(
            writer,
            normalidad,
            "11_normalidad",
            "Tests de normalidad - muestra total",
        )

        # 12. EDA descriptivo y conversiones
        fila = 0
        fila = escribir_tabla_excel(
            writer,
            merged.describe(include="all").transpose().reset_index().rename(columns={"index": "variable"}),
            "12_eda_descriptivo",
            "Resumen descriptivo general",
            startrow=fila,
        )
        for col in CONVERSION_COLS:
            fila = escribir_tabla_excel(
                writer,
                tabla_conversion(merged, col, min_n=100),
                "12_eda_descriptivo",
                f"Tasa de conversión por {col}",
                startrow=fila,
            )

        crear_hoja_indice(writer, nombre_archivo="Excel final limpio de resultados")

    print(f"Excel final limpio creado en: {ruta}")
